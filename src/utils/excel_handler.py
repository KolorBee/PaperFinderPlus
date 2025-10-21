"""
Excel文件输出处理模块，专门用于生成Excel格式的结果文件
"""
import os
import re

# 检查是否安装了pandas和openpyxl
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False
    print("警告：未安装pandas或openpyxl库，无法输出Excel格式文件")
except ImportError:
    EXCEL_SUPPORTED = False
    print("警告：未安装pandas或openpyxl库，无法输出Excel格式文件")


def _ensure_remove_sheet(writer, output_file, sheet_name):
    """确保在 workbook 中移除指定名称的工作表，以便后续写入可以创建/替换它。

    writer: pandas ExcelWriter（可能包含 .book）
    output_file: Excel 文件路径（在 writer.book 不存在时可用于加载）
    sheet_name: 要移除的工作表名
    """
    try:
        book = getattr(writer, 'book', None)
        # 如果 writer 没有 book 属性但文件存在，尝试加载工作簿
        if book is None and os.path.exists(output_file):
            book = openpyxl.load_workbook(output_file)
            writer.book = book

        if book is None:
            return

        if sheet_name in book.sheetnames:
            # 直接从 workbook 中移除该 sheet
            sheet = book[sheet_name]
            book.remove(sheet)
            # 尝试同步 writer.sheets 映射（若存在）
            if hasattr(writer, 'sheets'):
                try:
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                except Exception:
                    pass
    except Exception:
        # 容错：如果移除失败也不要阻塞写入流程
        return

def check_excel_support():
    """检查是否支持Excel输出"""
    return EXCEL_SUPPORTED

def extract_doi_link(paper):
    """
    从论文条目中提取DOI链接
    
    Args:
        paper: 原始论文条目文本
        
    Returns:
        (paper_title, doi_link): 论文标题和DOI链接（无链接则返回空字符串）
    """
    # 检查是否包含标准DOI链接
    doi_match = re.search(r'\[DOI: (https?://doi\.org/[^\]]+)\]', paper)
    if doi_match:
        doi_link = doi_match.group(1)
        # 移除原始DOI标记
        paper_title = paper.replace(f" [DOI: {doi_link}]", "")
        return paper_title, doi_link
    
    # 检查是否包含会议链接作为DOI（新增支持）
    conf_match = re.search(r'\[DOI: (https?://[^\]]+)\]', paper)
    if conf_match:
        doi_link = conf_match.group(1)
        # 移除原始DOI标记
        paper_title = paper.replace(f" [DOI: {doi_link}]", "")
        return paper_title, doi_link
    
    # 如果使用其他格式，也尝试提取
    doi_match = re.search(r'DOI: (https?://doi\.org/\S+)', paper)
    if doi_match:
        doi_link = doi_match.group(1)
        # 移除原始DOI标记
        paper_title = paper.replace(f" DOI: {doi_link}", "")
        return paper_title, doi_link
    
    # 尝试提取其他格式的会议链接（新增支持）
    conf_match = re.search(r'DOI: (https?://\S+)', paper)
    if conf_match:
        doi_link = conf_match.group(1)
        # 移除原始DOI标记
        paper_title = paper.replace(f" DOI: {doi_link}", "")
        return paper_title, doi_link
    
    return paper, ""

def apply_excel_formatting(worksheet):
    """应用Excel格式：居中对齐并自动调整列宽"""
    if not EXCEL_SUPPORTED:
        return
    
    # 设置列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # 找出最长的单元格内容
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # 设置列宽度（根据内容动态调整，但有最小和最大值）
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 设置标题行粗体
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    
    # 设置所有单元格居中对齐
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

def save_topic_results(results, all_papers, topic_name, output_file, is_journal=True):
    """保存结果到Excel文件，排版更美观"""
    if not EXCEL_SUPPORTED:
        print(f"警告：未安装pandas或openpyxl库，无法生成Excel文件：{output_file}")
        return False
    
    try:
        # 检查文件是否已存在
        if os.path.exists(output_file):
            # 读取现有文件
            try:
                with pd.ExcelFile(output_file) as xls:
                    # 获取所有工作表
                    sheet_names = xls.sheet_names
                    
                    # 读取现有数据
                    existing_data = {}
                    for sheet in sheet_names:
                        existing_data[sheet] = pd.read_excel(output_file, sheet_name=sheet)
                    
                    # 处理要添加的数据
                    overview_data = []
                    if all_papers:
                        # 确定下一个序号
                        next_index = 1
                        if "论文总览" in sheet_names and not existing_data["论文总览"].empty:
                            try:
                                next_index = existing_data["论文总览"]["序号"].max() + 1
                            except:
                                next_index = len(existing_data["论文总览"]) + 1
                        
                        # 准备新数据
                        for idx, paper in enumerate(all_papers, next_index):
                            title, doi = extract_doi_link(paper)
                            # 提取会议/期刊信息(如有)
                            venue_info = ""
                            if title.startswith('['):
                                venue_end = title.find(']')
                                if venue_end > 0:
                                    venue_info = title[1:venue_end]
                                    title = title[venue_end+1:].strip()
                            
                            overview_data.append({
                                "序号": idx,
                                "论文标题": title,
                                "会议/期刊": venue_info,
                                "DOI链接": doi
                            })
                    
                    # 提取详细信息但合并到总览中
                    if results:
                        current_venue = ""
                        current_year = ""
                        
                        for line in results:
                            if line.startswith("## "):
                                # 新的会议/期刊部分
                                match = re.match(r"## (.*) (\d+)年", line)
                                if match:
                                    current_venue = match.group(1)
                                    current_year = match.group(2)
                    
                    # 创建ExcelWriter对象进行更新
                    mode_str = 'a' if openpyxl.__version__ >= '3.0.0' else 'w'
                    with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode_str) as writer:
                        # 对于旧版openpyxl，需要先加载文件
                        if openpyxl.__version__ < '3.0.0' and os.path.exists(output_file):
                            writer.book = openpyxl.load_workbook(output_file)
                            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

                        # 安全获取 writer.sheets（某些 pandas 版本下 writer 可能不公开该属性）
                        writer_sheets = getattr(writer, 'sheets', {})

                        # 更新总览工作表 - 只做追加，不考虑内容匹配
                        if "论文总览" in sheet_names:
                            # 合并现有数据和新数据，简单追加
                            current_data = existing_data["论文总览"]
                            new_data = pd.DataFrame(overview_data)
                            updated_data = pd.concat([current_data, new_data], ignore_index=True)

                            # 如果是追加模式，需要先移除已存在的 '论文总览'，避免 to_excel 抛出错误
                            if mode_str == 'a':
                                _ensure_remove_sheet(writer, output_file, "论文总览")

                            # 写入更新后的数据
                            updated_data.to_excel(writer, sheet_name="论文总览", index=False)
                        else:
                            # 创建新的总览工作表
                            pd.DataFrame(overview_data).to_excel(writer, sheet_name="论文总览", index=False)

                        # 保持其他工作表不变（如果有的话）
                        for sheet in sheet_names:
                            if sheet != "论文总览":
                                # 如果是追加模式，已有的工作表不需要再次写入
                                if not (mode_str == 'a' and sheet in writer_sheets):
                                    existing_data[sheet].to_excel(writer, sheet_name=sheet, index=False)
                
                    # 保存并应用格式
                    wb = openpyxl.load_workbook(output_file)
                    for sheet_name in wb.sheetnames:
                        apply_excel_formatting(wb[sheet_name])
                    wb.save(output_file)
                    
                    return True
            except Exception as e:
                print(f"更新Excel文件时出错: {e}")
                # 出错时创建新文件
                # 创建ExcelWriter对象
                writer = pd.ExcelWriter(output_file, engine='openpyxl')
        else:
            # 创建ExcelWriter对象
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            
        # 创建总览工作表，将DOI链接单独放在一列
        overview_data = []
        if all_papers:
            for idx, paper in enumerate(all_papers, 1):
                title, doi = extract_doi_link(paper)
                # 提取会议/期刊信息(如有)
                venue_info = ""
                if title.startswith('['):
                    venue_end = title.find(']')
                    if venue_end > 0:
                        venue_info = title[1:venue_end]
                        title = title[venue_end+1:].strip()
                
                overview_data.append({
                    "序号": idx,
                    "论文标题": title,
                    "会议/期刊": venue_info,
                    "DOI链接": doi
                })
        
        # 创建DataFrame并保存
        if overview_data:
            overview_df = pd.DataFrame(overview_data)
        else:
            overview_df = pd.DataFrame(columns=["序号", "论文标题", "会议/期刊", "DOI链接"])
        
        # 只保存总览工作表
        overview_df.to_excel(writer, sheet_name="论文总览", index=False)
        
        # 保存并应用格式
        writer.close()
        
        # 应用单元格格式
        wb = openpyxl.load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            apply_excel_formatting(wb[sheet_name])
        wb.save(output_file)
        
        return True
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")
        return False

def save_venue_result(venue_name, venue_full_name, year, papers, source_link, 
                     volume_link=None, contents_link=None, topic_name="", 
                     is_journal=True, output_file=None):
    """保存单个会议/期刊结果到Excel文件"""
    if not EXCEL_SUPPORTED:
        print(f"警告：未安装pandas或openpyxl库，无法生成Excel文件：{output_file}")
        return False
    
    try:
        # 检查文件是否存在
        if os.path.exists(output_file):
            # 读取现有的Excel文件
            try:
                with pd.ExcelFile(output_file) as xls:
                    # 获取所有工作表
                    sheet_names = xls.sheet_names
                    
                    # 读取现有数据
                    existing_data = {}
                    for sheet in sheet_names:
                        existing_data[sheet] = pd.read_excel(output_file, sheet_name=sheet)
                    
                    # 准备数据
                    venue_full_display = venue_name
                    if venue_full_name:
                        venue_full_display = f"{venue_name} ({venue_full_name})"
                    
                    # 准备新数据 - 将DOI链接单独放在一列
                    new_overview_rows = []
                    
                    # 确定总览表的下一个序号
                    next_index = 1
                    if "论文总览" in sheet_names and not existing_data["论文总览"].empty:
                        try:
                            next_index = existing_data["论文总览"]["序号"].max() + 1
                        except:
                            next_index = len(existing_data["论文总览"]) + 1
                    
                    for paper_idx, paper in enumerate(papers):
                        title, doi = extract_doi_link(paper)
                        
                        # 总览 - 使用递增的序号和完整会议名称
                        new_overview_rows.append({
                            "序号": next_index + paper_idx,
                            "论文标题": title,
                            "会议/期刊": f"{venue_full_display} {year}",
                            "DOI链接": doi
                        })
                    
                    # 创建新的Excel写入器
                    mode_str = 'a' if openpyxl.__version__ >= '3.0.0' else 'w'
                    with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode_str) as writer:
                        # 对于旧版openpyxl，需要先删除文件
                        if openpyxl.__version__ < '3.0.0' and os.path.exists(output_file):
                            writer.book = openpyxl.load_workbook(output_file)
                            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

                        writer_sheets = getattr(writer, 'sheets', {})

                        # 写入总览工作表
                        if "论文总览" in sheet_names:
                            # 更新总览 - 简单追加数据，不考虑内容匹配
                            current_data = existing_data["论文总览"]
                            new_data = pd.DataFrame(new_overview_rows)
                            # 直接追加，不检查重复
                            updated_data = pd.concat([current_data, new_data], ignore_index=True)

                            # 如果是追加模式，需要先移除已存在的 '论文总览'，避免 to_excel 抛出错误
                            if mode_str == 'a':
                                _ensure_remove_sheet(writer, output_file, "论文总览")

                            updated_data.to_excel(writer, sheet_name="论文总览", index=False)
                        else:
                            # 创建新的总览工作表
                            pd.DataFrame(new_overview_rows).to_excel(writer, sheet_name="论文总览", index=False)

                        # 保持其他工作表不变（如果有的话）
                        for sheet in sheet_names:
                            if sheet != "论文总览":
                                # 如果是追加模式，已有的工作表不需要再次写入
                                if not (mode_str == 'a' and sheet in writer_sheets):
                                    existing_data[sheet].to_excel(writer, sheet_name=sheet, index=False)
                    
                    # 保存Excel后应用格式
                    wb = openpyxl.load_workbook(output_file)
                    for sheet_name in wb.sheetnames:
                        apply_excel_formatting(wb[sheet_name])
                    wb.save(output_file)
                    
                    return True
            except Exception as e:
                print(f"更新Excel文件时出错: {e}")
                # 创建新文件
                return create_new_excel_file(output_file, venue_name, venue_full_name, year, 
                                         papers, source_link, volume_link, contents_link, is_journal)
        else:
            # 创建新的Excel文件
            return create_new_excel_file(output_file, venue_name, venue_full_name, year, 
                                     papers, source_link, volume_link, contents_link, is_journal)
    except Exception as e:
        print(f"处理Excel文件时出错: {e}")
        return False

def create_new_excel_file(output_file, venue_name, venue_full_name, year, papers, 
                        source_link, volume_link=None, contents_link=None, is_journal=True):
    """创建新的Excel文件"""
    if not EXCEL_SUPPORTED:
        print(f"警告：未安装pandas或openpyxl库，无法生成Excel文件：{output_file}")
        return False
    
    try:
        # 创建ExcelWriter对象
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # 准备数据
        venue_full_display = venue_name
        if venue_full_name:
            venue_full_display = f"{venue_name} ({venue_full_name})"
        
        # 提取数据，将DOI单独放在一列
        overview_data = []
        
        for idx, paper in enumerate(papers, 1):
            title, doi = extract_doi_link(paper)
            
            # 总览数据 - 包含完整会议名称和所有信息
            overview_data.append({
                "序号": idx,
                "论文标题": title,
                "会议/期刊": f"{venue_full_display} {year}",
                "DOI链接": doi
            })
        
        # 创建DataFrames
        overview_df = pd.DataFrame(overview_data)
        
        # 只保存总览工作表
        overview_df.to_excel(writer, sheet_name="论文总览", index=False)
        
        # 保存Excel文件
        writer.close()
        
        # 应用格式
        wb = openpyxl.load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            apply_excel_formatting(wb[sheet_name])
        wb.save(output_file)
        
        return True
    except Exception as e:
        print(f"创建Excel文件时出错: {e}")
        return False 